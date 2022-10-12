#!/usr/bin/python3
import openpyxl

wb = openpyxl.load_workbook("student.xlsx")
ws = wb['Sheet1']
row_id = 1
for row in ws:
	sum_v = 0
	if row_id != 1:
		sum_v += ws.cell(row = row_id, column = 3).value * 0.3
		sum_v += ws.cell(row = row_id, column = 4).value * 0.35
		sum_v += ws.cell(row = row_id, column = 5).value * 0.34
		sum_v += ws.cell(row = row_id, column = 6).value
		ws.cell(row = row_id, column = 7).value = sum_v
	row_id += 1
# grade
student_num = row_id - 1
ws = wb['Sheet1']
row_id1 = 1
count = []
for i in range(student_num + 1):
	count.append(0)
for row in ws:
	row_id2 = 1
	if row_id1 != 1:
		rank = 1
		for row in ws:
			if row_id2 != 1:
				if ws.cell(row = row_id1, column = 7).value < ws.cell(row = row_id2, column = 7).value:
					rank += 1
				if ws.cell(row = row_id1, column = 7).value == ws.cell(row = row_id2, column = 7).value and row_id1 != row_id2:
					count[row_id1] += 1
			row_id2 += 1
		ws.cell(row = row_id1, column = 8).value = rank
	row_id1 += 1

row_id3 = 1
for row in ws:
	if row_id3 != 1:
		if ws.cell(row = row_id3, column = 8).value + count[row_id3] <= student_num * 0.15 - count[row_id3]:
			ws.cell(row = row_id3, column = 8).value = 'A+'
		elif ws.cell(row = row_id3, column = 8).value + count[row_id3] <= student_num * 0.3 - count[row_id3]:
			ws.cell(row = row_id3, column = 8).value = 'A'
		elif ws.cell(row = row_id3, column = 8).value + count[row_id3] <= student_num * 0.5 - count[row_id3]:
			ws.cell(row = row_id3, column = 8).value = 'B+'
		elif ws.cell(row = row_id3, column = 8).value + count[row_id3] <= student_num * 0.7 - count[row_id3]:
			ws.cell(row = row_id3, column = 8).value = 'B'
		elif ws.cell(row = row_id3, column = 8).value + count[row_id3] <= student_num * 0.85 - count[row_id3]:
			ws.cell(row = row_id3, column = 8).value = 'C+'
		else:
			ws.cell(row = row_id3, column = 8).value = 'C'
	row_id3 += 1

wb.save("student.xlsx")
 
